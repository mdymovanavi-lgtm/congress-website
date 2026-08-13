# -*- coding: utf-8 -*-
"""Collect site photos from original sources (not webp) and refresh фото-сайта.xlsx."""
from __future__ import annotations

import hashlib
import re
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

ROOT = Path(__file__).resolve().parent.parent
SOURCE_ROOT = Path(r"C:\Users\MSI-2\Desktop\Сайт\Контент\Фото")
OUT_DIR = ROOT / "фото-сайта"
XLSX = ROOT / "фото-сайта.xlsx"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif", ".jfif"}
SKIP_PREFIXES = ("assets/brand/", "assets/files/")

RK = "фотобанк РК"
OURS = "наши фото"

RK_HINTS = (
    "фотобанк",
    "фото из вк",
    "группа вк",
    "источник неизвест",
    "заставки проектов на глвное от риа",
    "фотобанк рк",
)
OURS_HINTS = (
    "гугл диск",
    "наши фото",
    "наше из",
    "музей фаберже",
    "дем. конференция",
    "демограф",
)

PAGE_LABELS = {
    "index.html": "Главная",
    "about.html": "О нас",
    "services.html": "Услуги",
    "volunteers.html": "Волонтёрам",
    "services/event.html": "Услуга — Мероприятие под ключ",
    "services/staff.html": "Услуга — Подбор временного персонала",
    "services/service.html": "Услуга — Сервисное обеспечение",
    "services/accreditation.html": "Услуга — Аккредитация",
    "services/expo.html": "Услуга — Организация выставки",
    "services/program.html": "Услуга — Деловая программа",
    "services/navigation.html": "Услуга — Навигация",
    "services/participants.html": "Услуга — Работа с участниками",
    "projects/pmef.html": "Проект — ПМЭФ",
    "projects/vef.html": "Проект — ВЭФ",
    "projects/russia-africa.html": "Проект — Саммит Россия — Африка",
    "projects/vfm.html": "Проект — ВФМ",
    "projects/brics.html": "Проект — Саммит БРИКС",
    "projects/demograph.html": "Проект — Демография ДВ",
    "projects/buddhist.html": "Проект — Буддийский форум",
    "projects/rus-china.html": "Проект — Российско-китайский форум",
    "calendar/vef-2026.html": "Календарь — ВЭФ-2026",
    "calendar/fok-2026.html": "Календарь — ФОК-2026",
    "calendar/ren-2026.html": "Календарь — РЭН-2026",
    "calendar/kmu-2026.html": "Календарь — КМУ-2026",
    "script.js": "Волонтёрам (календарь)",
}

RIA_FILES = {
    "pmef.jpeg": SOURCE_ROOT / "Заставки проектов на глвное от РИА" / "ПМЭФ.jpeg",
    "vef.jpeg": SOURCE_ROOT / "Заставки проектов на глвное от РИА" / "ВЭФ.jpeg",
    "russia-africa.jpeg": SOURCE_ROOT / "Заставки проектов на глвное от РИА" / "Россия-Африка.jpeg",
    "vfm.jpeg": SOURCE_ROOT / "Заставки проектов на глвное от РИА" / "ВСемирный фестиваль молодежи.jpeg",
}

SIDECAR = {
    "svc-personnel.webp": "svc-personnel-source.jpg",
    "svc-participants.webp": "svc-participants-source.jpg",
    "svc-event.webp": "svc-event-faberge.jpg",
}

SERVICE_FALLBACKS = {
    "svc-navigation.webp": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "IMG_2180.jpg",
    "svc-service.webp": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "CHER4518.jpg",
    "svc-program.webp": SOURCE_ROOT / "Фото проектов" / "Дем. конференция" / "DSC_1998.jpg",
}

NOTES_PROJECTS = {
    "pmef": "ПМЭФ",
    "vef": "ВЭФ",
    "russia-africa": "Р-А",
    "brics": "БРИКС",
    "vfm": "ВФМ",
    "demograph": "Дем конференция",
    "buddhist": "МБФ",
    "rus-china": "РКФ",
}

BLOCK_COLUMNS = {
    "hero": 1,
    "staff": 2,
    "service": 3,
    "navigation": 4,
    "participants": 5,
    "exhibition": 6,
}

BLOCK_OVERRIDES: dict[tuple[str, str], Path] = {
    ("pmef", "hero"): SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "APR_1513.jpg",
    ("pmef", "staff"): SOURCE_ROOT / "ПМЭФ" / "ПМЭФ (группа ВК)" / "5fc9b-qhOPjvXE66-IqWqZVUlm1248bmWf7K7CTt3B5WBYc12LRMmRUJVhiEdBGWpuLUJa6H564iiJM0Pf2w9tYV.jpg",
    ("pmef", "service"): SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "CHER4518.jpg",
    ("pmef", "navigation"): SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "IMG_2180.jpg",
    ("pmef", "participants"): SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "CHER4674.jpg",
    ("vef", "hero"): SOURCE_ROOT / "ВЭФ" / "ВЭФ фотобанк РК" / "ВЭФ ФБ 2.jpg",
    ("vef", "staff"): SOURCE_ROOT / "ВЭФ" / "ВЭФ_фото из ВК" / "ovCMaA-_pTvy17HROu31pY6hLaPySXrkqNeOyeiHqSaPwbJ50KB3Z__PWrymXSehq76UiYTPG1M7OE1yIFPPS5j8.jpg",
    ("vef", "service"): SOURCE_ROOT / "ВЭФ" / "ВЭФ_фото из ВК" / "glW8m5sHpkx6wzMZvzbuLzOXpSQX1c7U6Ty68btcYVM6ysP9c2WVndNmxrAeBF7ORA88vNl6ylSUGiPUnf4RT6M6.jpg",
    ("vef", "navigation"): SOURCE_ROOT / "ВЭФ" / "ВЭФ (источник неизвестный)" / "photo_2026-03-10_13-13-42.jpg",
    ("russia-africa", "hero"): SOURCE_ROOT / "Р-А" / "Р-А (фотобанк)" / "RS_08644.jpg",
    ("russia-africa", "staff"): SOURCE_ROOT / "Р-А" / "Р-А (группа вк)" / "ZBHyAc8YR8lCHIq-9leZ6w2o68EmNQR5JV_Q7Y1BhpI32BXLwzyScPxjT74T-kkvaifvmay-.jpg",
    ("russia-africa", "service"): SOURCE_ROOT / "Р-А" / "Р-А (группа вк)" / "in-OgLvJOmnifIYS4o9N6oKxw41up2g7LgrK8N7SS6mbGF_Z9R2qnJn6iDmDVUrmAdCyu7SuCmGUeov4MVZDOGgU.jpg",
    ("russia-africa", "navigation"): SOURCE_ROOT / "Р-А" / "Р-А (группа вк)" / "Gf96T2M5qJFjuTDQ8dzrlFu5S3F5RBWvmToDexjl2C-FjoTS5nsbztNUhJBg30sweJIAWx8p.jpg",
    ("brics", "service"): SOURCE_ROOT / "Музей фаберже" / "photo_2026-08-13_14-00-42.jpg",
}

PMEF_GALLERY = {
    "01": SOURCE_ROOT / "ПМЭФ" / "ПМЭФ (группа ВК)" / "QvGvD5k440mQbgpwALoycjDEH4VISOQTq4j4P0_bTYvRtk3zEE57Z3nBff3W8mmFGY5hXI44rxpPYONIJDSSdPTe.jpg",
    "02": SOURCE_ROOT / "ПМЭФ" / "ПМЭФ (группа ВК)" / "XIGWgz56qEPhwVekhHNQblCPBYA0939ZXw1ke_woLw1JBh6xaV5dhBwVBnnO_C3l3JLAB08b3Ym8kGEs-yF3oFWW.jpg",
    "03": SOURCE_ROOT / "ПМЭФ" / "ПМЭФ (группа ВК)" / "m7P3nDK-b5NFint1cAe5euS_5YvbDlylvHMusULOAwARHw4Rxo9ZqQyMf4wqAhl0fQhrOpG2lNqYYaDaxqNEPV4t.jpg",
    "04": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "CHER4518.jpg",
    "05": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "IMG_2180.jpg",
    "06": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "NAV_6946.jpg",
    "07": SOURCE_ROOT / "ПМЭФ" / "Фотобанк РК" / "ПМЭФ VPEV9989.jpg",
}

PROJECT_SOURCE_DIRS = {
    "pmef": [
        SOURCE_ROOT / "ПМЭФ",
    ],
    "vef": [
        SOURCE_ROOT / "ВЭФ",
    ],
    "russia-africa": [
        SOURCE_ROOT / "Р-А",
    ],
    "brics": [
        SOURCE_ROOT / "БРИКС",
    ],
    "vfm": [
        SOURCE_ROOT / "ВФМ" / "ВФМ (наши фото)",
        SOURCE_ROOT / "Фото проектов" / "4. ВФМ",
    ],
    "demograph": [
        SOURCE_ROOT / "Фото проектов" / "Дем. конференция",
    ],
}


def file_hash(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_rel(path: str) -> str:
    path = path.strip().strip('"').replace("\\", "/")
    while path.startswith("../"):
        path = path[3:]
    return path


def page_label(rel_html: str) -> str:
    key = rel_html.replace("\\", "/").removesuffix(" (js)")
    return PAGE_LABELS.get(key, key)


def classify_source_path(path: Path) -> str:
    joined = " / ".join(p.lower() for p in path.parts)
    if any(h in joined for h in OURS_HINTS):
        return OURS
    if any(h in joined for h in RK_HINTS):
        return RK
    return RK


def collect_site_refs() -> dict[str, set[str]]:
    refs: dict[str, set[str]] = {}
    patterns = [
        re.compile(r'(?:src|poster|href)=["\']([^"\']+\.(?:jpg|jpeg|png|webp|gif|avif))["\']', re.I),
        re.compile(r"""img:\s*['"]([^'"]+\.(?:jpg|jpeg|png|webp|gif))['"]""", re.I),
    ]
    for fp in ROOT.rglob("*"):
        if fp.suffix.lower() not in {".html", ".css", ".js"}:
            continue
        if "_archive" in fp.parts:
            continue
        text = fp.read_text(encoding="utf-8", errors="ignore")
        page = str(fp.relative_to(ROOT)).replace("\\", "/")
        for pat in patterns:
            for m in pat.finditer(text):
                rel = normalize_rel(m.group(1))
                if not rel.startswith("assets/"):
                    continue
                if any(rel.startswith(p) for p in SKIP_PREFIXES):
                    continue
                if "/clients/" in rel:
                    continue
                refs.setdefault(rel, set()).add(page)
    return refs


def build_hash_index(folder: Path) -> dict[str, Path]:
    index: dict[str, Path] = {}
    if not folder.exists():
        return index
    for fp in folder.rglob("*"):
        if fp.is_file() and fp.suffix.lower() in IMAGE_EXTS and fp.suffix.lower() != ".webp":
            try:
                index[file_hash(fp)] = fp
            except OSError:
                pass
    return index


def build_name_index(folder: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    for fp in folder.rglob("*"):
        if fp.is_file() and fp.suffix.lower() in IMAGE_EXTS and fp.suffix.lower() != ".webp":
            index.setdefault(fp.name.lower(), []).append(fp)
            index.setdefault(fp.stem.lower(), []).append(fp)
    return index


def normalize_notes_path(path: Path) -> Path:
    if path.exists():
        return path
    parts = path.parts
    if "Фото" in parts:
        idx = parts.index("Фото")
        candidate = SOURCE_ROOT.joinpath(*parts[idx + 1 :])
        if candidate.exists():
            return candidate
    return path


def load_notes_blocks() -> dict[tuple[str, str], Path]:
    blocks: dict[tuple[str, str], Path] = {}
    if not XLSX.exists():
        return blocks
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "мои заметки" not in wb.sheetnames:
        return blocks
    ws = wb["мои заметки"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return blocks
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        project_name = str(row[0]).strip()
        project_key = next((k for k, v in NOTES_PROJECTS.items() if v == project_name), None)
        if not project_key:
            continue
        for block, col in BLOCK_COLUMNS.items():
            if col >= len(row):
                continue
            cell = row[col]
            if not cell or str(cell).strip() in {"", "-", " - "}:
                continue
            path = normalize_notes_path(Path(str(cell).strip().strip('"')))
            if path.exists():
                blocks[(project_key, block)] = path
    return blocks


def preprocess_like_site(path: Path) -> Image.Image:
    im = ImageOps.exif_transpose(Image.open(path)).convert("RGB")
    w, h = im.size
    scale = min(1.0, 1920 / max(w, h))
    if scale < 1:
        im = im.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    return im


def average_hash(path: Path, size: int = 16) -> int:
    im = preprocess_like_site(path).convert("L").resize((size, size), Image.Resampling.LANCZOS)
    pixels = list(im.getdata())
    avg = sum(pixels) / len(pixels)
    bits = 0
    for i, p in enumerate(pixels):
        if p >= avg:
            bits |= 1 << i
    return bits


def hamming(a: int, b: int) -> int:
    return (a ^ b).bit_count()


def project_pool(project: str) -> list[Path]:
    pool: list[Path] = []
    for folder in PROJECT_SOURCE_DIRS.get(project, []):
        if not folder.exists():
            continue
        for fp in folder.rglob("*"):
            if fp.is_file() and fp.suffix.lower() in IMAGE_EXTS and fp.suffix.lower() != ".webp":
                pool.append(fp)
    return pool


def ra_gallery_source(index: str) -> Path | None:
    folder = SOURCE_ROOT / "Р-А" / "Р-А (группа вк)"
    if not folder.exists():
        return None
    files = sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTS and p.suffix.lower() != ".webp"],
        key=lambda p: p.name.lower(),
    )
    i = int(index) - 1
    if 0 <= i < len(files):
        return files[i]
    return None


def parse_project_and_block(site_rel: str) -> tuple[str | None, str | None, str | None]:
    rel = site_rel.replace("\\", "/").lower()
    m = re.search(r"projects/([a-z0-9-]+)/gallery/(\d+)\.", rel)
    if m:
        return m.group(1), "gallery", m.group(2)
    m = re.search(r"projects/([a-z0-9-]+)/([a-z0-9-]+)\.", rel)
    if m:
        return m.group(1), m.group(2), None
    m = re.search(r"projects/([a-z0-9-]+)\.", rel)
    if m:
        return m.group(1), "hero", None
    return None, None, None


def match_by_phash(site_path: Path, candidates: list[Path]) -> Path | None:
    if not candidates:
        return None
    try:
        target = average_hash(site_path)
    except OSError:
        return None
    best: Path | None = None
    best_score = 999
    for cand in candidates:
        try:
            score = hamming(target, average_hash(cand))
        except OSError:
            continue
        if score < best_score:
            best_score = score
            best = cand
    return best if best_score <= 12 else None


def resolve_source_file(
    site_rel: str,
    site_path: Path,
    src_hash: dict[str, Path],
    src_name: dict[str, list[Path]],
    notes_blocks: dict[tuple[str, str], Path],
) -> tuple[str, Path | None, str]:
    rel = site_rel.replace("\\", "/").lower()
    name = site_path.name.lower()

    def ok(p: Path) -> tuple[str, Path, str]:
        return classify_source_path(p), p, str(p)

    # Non-webp on site that already matches a source
    if site_path.suffix.lower() != ".webp":
        try:
            matched = src_hash.get(file_hash(site_path))
            if matched:
                return ok(matched)
        except OSError:
            pass

    # Sidecar originals
    sidecar_name = SIDECAR.get(name)
    if sidecar_name:
        sidecar = site_path.parent / sidecar_name
        if sidecar.exists():
            matched = src_hash.get(file_hash(sidecar))
            if matched:
                return ok(matched)

    # Service heroes without sidecar
    fallback = SERVICE_FALLBACKS.get(name)
    if fallback and fallback.exists():
        return ok(fallback)

    project, block, gallery_idx = parse_project_and_block(site_rel)

    # Explicit overrides and notes
    if project and block:
        override = BLOCK_OVERRIDES.get((project, block))
        if override and override.exists():
            return ok(override)
        noted = notes_blocks.get((project, block))
        if noted and noted.exists():
            return ok(noted)

    # PMEF gallery map
    if project == "pmef" and block == "gallery" and gallery_idx in PMEF_GALLERY:
        src = PMEF_GALLERY[gallery_idx]
        if src.exists():
            return ok(src)

    # Russia-Africa gallery: sorted VK folder
    if project == "russia-africa" and block == "gallery" and gallery_idx:
        src = ra_gallery_source(gallery_idx)
        if src:
            return ok(src)

    # RIA card thumbnails -> original jpeg cards
    if name in RIA_FILES:
        src = RIA_FILES[name]
        if src.exists():
            return ok(src)

    # Card webp like pmef.webp -> project hero source
    if project and block == "hero" and name.endswith(".webp"):
        hero_src = notes_blocks.get((project, "hero")) or BLOCK_OVERRIDES.get((project, "hero"))
        if hero_src and hero_src.exists():
            return ok(hero_src)

    # Demograph DSC
    stem = site_path.stem.lower()
    if "/demograph/" in rel or stem.startswith("dsc_"):
        demo = SOURCE_ROOT / "Фото проектов" / "Дем. конференция" / f"{stem}.jpg"
        if demo.exists():
            return ok(demo)

    # Stem lookup, prefer non-webp
    for candidate in src_name.get(stem, []):
        if candidate.suffix.lower() != ".webp":
            return ok(candidate)

    # Sibling JPG hash (processed copy on site)
    if site_path.suffix.lower() == ".webp":
        jpg = site_path.with_suffix(".jpg")
        if jpg.exists():
            matched = src_hash.get(file_hash(jpg))
            if matched:
                return ok(matched)

    # Visual match within project pool, then globally
    if project:
        matched = match_by_phash(site_path, project_pool(project))
        if matched:
            return ok(matched)

    global_pool = [
        fp
        for fp in SOURCE_ROOT.rglob("*")
        if fp.is_file() and fp.suffix.lower() in IMAGE_EXTS and fp.suffix.lower() != ".webp"
    ]
    matched = match_by_phash(site_path, global_pool)
    if matched:
        return ok(matched)

    if "hero-poster" in rel:
        return OURS, None, "Кадр из видео сайта (оригинал не в фотобанке)"

    return RK, None, "Исходник не найден"


def flat_name(site_rel: str, source: Path) -> str:
    rel = site_rel.replace("\\", "/")
    for prefix in ("assets/photo/", "assets/video/"):
        if rel.startswith(prefix):
            rel = rel[len(prefix):]
            break
    rel = rel.rsplit(".", 1)[0].replace("/", " — ")
    return f"{rel} — {source.name}"


def unique_dest(dest_dir: Path, name: str) -> Path:
    target = dest_dir / name
    if not target.exists():
        return target
    stem = Path(name).stem
    suffix = Path(name).suffix
    i = 2
    while True:
        candidate = dest_dir / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def copy_original(site_rel: str, category: str, source: Path) -> Path:
    dest_root = OUT_DIR / category
    dest_root.mkdir(parents=True, exist_ok=True)
    dest = unique_dest(dest_root, flat_name(site_rel, source))
    shutil.copy2(source, dest)
    return dest


def load_notes_sheet(wb_path: Path) -> list[list]:
    if not wb_path.exists():
        return []
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    if "мои заметки" not in wb.sheetnames:
        return []
    return [list(row) for row in wb["мои заметки"].iter_rows(values_only=True)]


def save_workbook(wb: Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(path.stem + "-обновлено.xlsx")
        wb.save(alt)
        return alt


def main() -> None:
    refs = collect_site_refs()
    src_hash = build_hash_index(SOURCE_ROOT)
    src_name = build_name_index(SOURCE_ROOT)
    notes_blocks = load_notes_blocks()
    notes = load_notes_sheet(XLSX)

    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    (OUT_DIR / RK).mkdir(parents=True)
    (OUT_DIR / OURS).mkdir(parents=True)

    rows: list[tuple[str, str, str]] = []
    copied = 0
    skipped = 0

    for site_rel in sorted(refs):
        if "/thanks/" in site_rel.replace("\\", "/").lower():
            continue

        site_path = ROOT / site_rel.replace("/", "\\")
        pages = ", ".join(sorted(page_label(p) for p in refs[site_rel]))

        if not site_path.exists():
            rows.append((pages, site_rel, "файл не найден на сайте"))
            continue

        category, source_file, source_label = resolve_source_file(
            site_rel, site_path, src_hash, src_name, notes_blocks
        )

        if source_file and source_file.exists() and source_file.suffix.lower() != ".webp":
            copy_original(site_rel, category, source_file)
            copied += 1
            rows.append((pages, site_rel.replace("\\", "/"), source_label))
        elif site_path.suffix.lower() != ".webp":
            copy_original(site_rel, category, site_path)
            copied += 1
            rows.append((pages, site_rel.replace("\\", "/"), str(site_path)))
        else:
            skipped += 1
            rows.append((pages, site_rel.replace("\\", "/"), source_label))

    wb = Workbook()
    ws = wb.active
    ws.title = "Фото на сайте"
    headers = ["Страница", "Ссылка на фото", "Источник"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append(list(row))

    widths = [44, 52, 72]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"

    if notes:
        dst = wb.create_sheet("мои заметки")
        for row in notes:
            dst.append(row)

    saved = save_workbook(wb, XLSX)

    rk_count = sum(1 for _ in (OUT_DIR / RK).rglob("*") if _.is_file())
    ours_count = sum(1 for _ in (OUT_DIR / OURS).rglob("*") if _.is_file())
    print(f"References on site: {len(refs)}")
    print(f"Copied originals: {copied}")
    print(f"Skipped (no source): {skipped}")
    print(f"{RK}: {rk_count} files")
    print(f"{OURS}: {ours_count} files")
    print(f"Excel: {len(rows)} rows -> {saved}")
    print(f"Folder: {OUT_DIR}")


if __name__ == "__main__":
    main()
