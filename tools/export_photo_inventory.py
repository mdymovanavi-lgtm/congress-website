# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "фото-сайта.xlsx"

ROWS = [
    ("ГЛАВНАЯ", "HERO", "Обложка фонового видео", "assets/video/hero-poster.jpg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.jpeg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка проекта ВЭФ", "assets/photo/projects/vef.jpeg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка саммита Россия — Африка", "assets/photo/projects/russia-africa.jpeg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка конференции «Демографическое развитие Дальнего Востока»", "assets/photo/projects/demograph/DSC_1998.jpg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка Всемирного фестиваля молодёжи", "assets/photo/projects/vfm.jpeg"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка Международного буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка Российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("ГЛАВНАЯ", "ПРОЕКТЫ", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности ФБТ-2026", "assets/photo/thanks/congress-фбт-2026.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности ПМЮФ 2025", "assets/photo/thanks/congress-пмюф-2025.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности РСД 2025", "assets/photo/thanks/congress-рсд-2025.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью диплома ВФМС 2017", "assets/photo/thanks/congress-вфмс-2017.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности АЭФ 2023", "assets/photo/thanks/аэф-2023-congress.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности губернатора Мурманской области 2023", "assets/photo/thanks/галкина-благодарность-от-губернатора-мурманской-области-2023.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности Минприроды 2023", "assets/photo/thanks/галкина-благодарность-от-министерства-природных-ресурсов-2023.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности КМУ 2023", "assets/photo/thanks/кму-2023-congress.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности АС 2019", "assets/photo/thanks/congress-ас-2019.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности Росатом", "assets/photo/thanks/congress-благодарность-росатом-congress.webp"),
    ("ГЛАВНАЯ", "БЛАГОДАРНОСТИ", "Превью благодарности ВНОТ 2023", "assets/photo/thanks/congress-внот-2023.webp"),
    ("УСЛУГИ", "Подбор временного персонала", "Фото временного персонала", "assets/photo/svc-personnel.webp"),
    ("УСЛУГИ", "Навигация", "Фото навигации", "assets/photo/svc-navigation.webp"),
    ("УСЛУГА — Мероприятие под ключ", "HERO", "Фото мероприятия под ключ", "assets/photo/svc-event.webp"),
    ("УСЛУГА — Мероприятие под ключ", "ПРОЕКТЫ", "Заставка конференции «Демографическое развитие Дальнего Востока»", "assets/photo/projects/demograph.webp"),
    ("УСЛУГА — Мероприятие под ключ", "ПРОЕКТЫ", "Заставка Российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("УСЛУГА — Мероприятие под ключ", "ПРОЕКТЫ", "Заставка Международного буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("УСЛУГА — Подбор временного персонала", "HERO", "Фото временного персонала", "assets/photo/svc-personnel.webp"),
    ("УСЛУГА — Подбор временного персонала", "ПРОЕКТЫ", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.webp"),
    ("УСЛУГА — Подбор временного персонала", "ПРОЕКТЫ", "Заставка проекта ВЭФ", "assets/photo/projects/vef.webp"),
    ("УСЛУГА — Подбор временного персонала", "ПРОЕКТЫ", "Заставка саммита Россия — Африка", "assets/photo/projects/russia-africa.webp"),
    ("УСЛУГА — Подбор временного персонала", "ПРОЕКТЫ", "Заставка ВФМ", "assets/photo/projects/vfm.webp"),
    ("УСЛУГА — Подбор временного персонала", "ПРОЕКТЫ", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("УСЛУГА — Сервисное обеспечение", "HERO", "Фото сервисного обеспечения", "assets/photo/svc-service.webp"),
    ("УСЛУГА — Сервисное обеспечение", "ПРОЕКТЫ", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.webp"),
    ("УСЛУГА — Сервисное обеспечение", "ПРОЕКТЫ", "Заставка проекта ВЭФ", "assets/photo/projects/vef.webp"),
    ("УСЛУГА — Сервисное обеспечение", "ПРОЕКТЫ", "Заставка саммита Россия — Африка", "assets/photo/projects/russia-africa.webp"),
    ("УСЛУГА — Сервисное обеспечение", "ПРОЕКТЫ", "Заставка ВФМ", "assets/photo/projects/vfm.webp"),
    ("УСЛУГА — Сервисное обеспечение", "ПРОЕКТЫ", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("УСЛУГА — Аккредитация", "HERO", "Фото аккредитации (конференция по демографии)", "assets/photo/projects/demograph.webp"),
    ("УСЛУГА — Аккредитация", "ПРОЕКТЫ", "Заставка конференции «Демографическое развитие Дальнего Востока»", "assets/photo/projects/demograph.webp"),
    ("УСЛУГА — Аккредитация", "ПРОЕКТЫ", "Заставка Российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("УСЛУГА — Аккредитация", "ПРОЕКТЫ", "Заставка Международного буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("УСЛУГА — Организация выставки", "HERO", "Фото выставочной застройки ВЭФ", "assets/photo/projects/vef.webp"),
    ("УСЛУГА — Организация выставки", "ПРОЕКТЫ", "Заставка проекта ВЭФ", "assets/photo/projects/vef.webp"),
    ("УСЛУГА — Организация выставки", "ПРОЕКТЫ", "Заставка Международного буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("УСЛУГА — Деловая программа", "HERO", "Фото деловой программы", "assets/photo/svc-program.webp"),
    ("УСЛУГА — Деловая программа", "ПРОЕКТЫ", "Заставка конференции «Демографическое развитие Дальнего Востока»", "assets/photo/projects/demograph.webp"),
    ("УСЛУГА — Деловая программа", "ПРОЕКТЫ", "Заставка Российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("УСЛУГА — Деловая программа", "ПРОЕКТЫ", "Заставка Международного буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("УСЛУГА — Навигация", "HERO", "Фото навигации", "assets/photo/svc-navigation.webp"),
    ("УСЛУГА — Навигация", "ПРОЕКТЫ", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.webp"),
    ("УСЛУГА — Навигация", "ПРОЕКТЫ", "Заставка проекта ВЭФ", "assets/photo/projects/vef.webp"),
    ("УСЛУГА — Навигация", "ПРОЕКТЫ", "Заставка саммита Россия — Африка", "assets/photo/projects/russia-africa.webp"),
    ("УСЛУГА — Навигация", "ПРОЕКТЫ", "Заставка ВФМ", "assets/photo/projects/vfm.webp"),
    ("УСЛУГА — Навигация", "ПРОЕКТЫ", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("УСЛУГА — Работа с участниками", "HERO", "Фото работы с участниками (саммит БРИКС)", "assets/photo/projects/brics.webp"),
    ("УСЛУГА — Работа с участниками", "ПРОЕКТЫ", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.webp"),
    ("УСЛУГА — Работа с участниками", "ПРОЕКТЫ", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("ПРОЕКТ — ПМЭФ", "HERO", "Заставка проекта ПМЭФ", "assets/photo/projects/pmef.jpeg"),
    ("ПРОЕКТ — ПМЭФ", "ЧТО ДЕЛАЛИ", "Фото персонала на ПМЭФ", "assets/photo/svc-personnel.webp"),
    ("ПРОЕКТ — ПМЭФ", "ЧТО ДЕЛАЛИ", "Фото сервисного обеспечения ПМЭФ", "assets/photo/svc-service.webp"),
    ("ПРОЕКТ — ПМЭФ", "ЧТО ДЕЛАЛИ", "Фото навигации ПМЭФ", "assets/photo/svc-navigation.webp"),
    ("ПРОЕКТ — ПМЭФ", "ЧТО ДЕЛАЛИ", "Фото работы с участниками ПМЭФ", "assets/photo/projects/pmef.jpeg"),
    ("ПРОЕКТ — ПМЭФ", "ФОТОГАЛЕРЕЯ", "Фото с ПМЭФ", "assets/photo/projects/pmef.webp"),
    ("ПРОЕКТ — ВЭФ", "HERO", "Заставка проекта ВЭФ", "assets/photo/projects/vef.jpeg"),
    ("ПРОЕКТ — ВЭФ", "ЧТО ДЕЛАЛИ", "Фото персонала на ВЭФ", "assets/photo/svc-personnel.webp"),
    ("ПРОЕКТ — ВЭФ", "ЧТО ДЕЛАЛИ", "Фото сервисного обеспечения ВЭФ", "assets/photo/svc-service.webp"),
    ("ПРОЕКТ — ВЭФ", "ЧТО ДЕЛАЛИ", "Фото навигации ВЭФ", "assets/photo/svc-navigation.webp"),
    ("ПРОЕКТ — ВЭФ", "ЧТО ДЕЛАЛИ", "Фото выставочной застройки ВЭФ", "assets/photo/projects/vef.jpeg"),
    ("ПРОЕКТ — ВЭФ", "ФОТОГАЛЕРЕЯ", "Фото с ВЭФ", "assets/photo/projects/vef.webp"),
    ("ПРОЕКТ — Саммит Россия — Африка", "HERO", "Заставка саммита Россия — Африка", "assets/photo/projects/russia-africa.jpeg"),
    ("ПРОЕКТ — Саммит Россия — Африка", "ЧТО ДЕЛАЛИ", "Фото персонала на саммите", "assets/photo/svc-personnel.webp"),
    ("ПРОЕКТ — Саммит Россия — Африка", "ЧТО ДЕЛАЛИ", "Фото сервисного обеспечения саммита", "assets/photo/svc-service.webp"),
    ("ПРОЕКТ — Саммит Россия — Африка", "ЧТО ДЕЛАЛИ", "Фото навигации саммита", "assets/photo/svc-navigation.webp"),
    ("ПРОЕКТ — Саммит Россия — Африка", "ФОТОГАЛЕРЕЯ", "Фото с саммита", "assets/photo/projects/russia-africa.webp"),
    ("ПРОЕКТ — ВФМ", "HERO", "Заставка ВФМ", "assets/photo/projects/vfm.jpeg"),
    ("ПРОЕКТ — ВФМ", "ЧТО ДЕЛАЛИ", "Фото персонала на ВФМ", "assets/photo/svc-personnel.webp"),
    ("ПРОЕКТ — ВФМ", "ЧТО ДЕЛАЛИ", "Фото сервисного обеспечения ВФМ", "assets/photo/svc-service.webp"),
    ("ПРОЕКТ — ВФМ", "ЧТО ДЕЛАЛИ", "Фото навигации ВФМ", "assets/photo/svc-navigation.webp"),
    ("ПРОЕКТ — ВФМ", "ЧТО ДЕЛАЛИ", "Фото экипировочного центра ВФМ", "assets/photo/projects/vfm.jpeg"),
    ("ПРОЕКТ — ВФМ", "ФОТОГАЛЕРЕЯ", "Фото с ВФМ", "assets/photo/projects/vfm.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "HERO", "Заставка саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "ЧТО ДЕЛАЛИ", "Фото персонала на саммите БРИКС", "assets/photo/svc-personnel.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "ЧТО ДЕЛАЛИ", "Фото сервисного обеспечения БРИКС", "assets/photo/svc-service.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "ЧТО ДЕЛАЛИ", "Фото навигации БРИКС", "assets/photo/svc-navigation.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "ЧТО ДЕЛАЛИ", "Фото работы с участниками БРИКС", "assets/photo/projects/brics.webp"),
    ("ПРОЕКТ — Саммит БРИКС", "ФОТОГАЛЕРЕЯ", "Фото с саммита БРИКС", "assets/photo/projects/brics.webp"),
    ("ПРОЕКТ — Демография ДВ", "HERO", "Заставка конференции", "assets/photo/projects/demograph/DSC_1998.jpg"),
    ("ПРОЕКТ — Демография ДВ", "ЧТО ДЕЛАЛИ", "Фото мероприятия под ключ (конференция)", "assets/photo/projects/demograph/DSC_1998.jpg"),
    ("ПРОЕКТ — Демография ДВ", "ФОТОГАЛЕРЕЯ", "Фото с конференции (31 фото в карусели)", "assets/photo/projects/demograph/*.webp"),
    ("ПРОЕКТ — Буддийский форум", "HERO", "Заставка буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("ПРОЕКТ — Буддийский форум", "ЧТО ДЕЛАЛИ", "Фото мероприятия под ключ (буддийский форум)", "assets/photo/projects/buddhist.webp"),
    ("ПРОЕКТ — Буддийский форум", "ФОТОГАЛЕРЕЯ", "Фото с буддийского форума", "assets/photo/projects/buddhist.webp"),
    ("ПРОЕКТ — Российско-китайский форум", "HERO", "Заставка российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("ПРОЕКТ — Российско-китайский форум", "ЧТО ДЕЛАЛИ", "Фото мероприятия под ключ (российско-китайский форум)", "assets/photo/projects/rus-china.webp"),
    ("ПРОЕКТ — Российско-китайский форум", "ФОТОГАЛЕРЕЯ", "Фото с российско-китайского форума", "assets/photo/projects/rus-china.webp"),
    ("ВОЛОНТЁРАМ", "HERO", "Фото волонтёров на ПМЭФ", "assets/photo/volunteers/hero.jpg"),
    ("ВОЛОНТЁРАМ", "КАЛЕНДАРЬ СОБЫТИЙ", "Заставка ПМЭФ-2026 (крупная карточка)", "assets/photo/projects/pmef.webp"),
    ("ВОЛОНТЁРАМ", "КАЛЕНДАРЬ СОБЫТИЙ", "Заставка ВЭФ-2026 (в списке событий)", "assets/photo/projects/vef.webp"),
    ("ВОЛОНТЁРАМ", "ФОТООТЧЁТЫ", "Заставка фотоотчёта ПМЮФ 2026", "assets/photo/volunteers/pmuf-2026.jpg"),
    ("ВОЛОНТЁРАМ", "ФОТООТЧЁТЫ", "Заставка фотоотчёта «Россия — АСЕАН»", "assets/photo/volunteers/russia-asean.jpg"),
    ("ВОЛОНТЁРАМ", "ФОТООТЧЁТЫ", "Заставка фотоотчёта ПМЭФ 2026", "assets/photo/volunteers/pmef-2026.jpg"),
    ("ВОЛОНТЁРАМ", "ФОТООТЧЁТЫ", "Заставка фотоотчёта КИФ 2026", "assets/photo/volunteers/kif-2026.jpg"),
    ("ВОЛОНТЁРАМ", "ФОТООТЧЁТЫ", "Заставка фотоотчёта МТЛФ 2026", "assets/photo/volunteers/mtlf-2026.jpg"),
    ("КАЛЕНДАРЬ — ВЭФ-2026", "HERO", "Заставка ВЭФ-2026", "assets/photo/projects/vef.webp"),
]

DEMOGRAPH_GALLERY = [
    "DSC_2928", "DSC_1911", "DSC_1967", "DSC_1998", "DSC_2077", "DSC_2089",
    "DSC_2216", "DSC_2333", "DSC_2338", "DSC_2356", "DSC_2376", "DSC_2433",
    "DSC_2454", "DSC_2858", "DSC_1906", "DSC_3442", "DSC_3458", "DSC_3687",
    "DSC_4052", "DSC_4314", "DSC_4346", "DSC_5160", "DSC_5161", "DSC_5294",
    "DSC_5315", "DSC_5318", "DSC_5344", "DSC_5359", "DSC_5524", "DSC_5560",
    "DSC_5565",
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Фото на сайте"

    headers = ["Страница", "Блок", "Описание фото", "Файл"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ROWS:
        if row[3] == "assets/photo/projects/demograph/*.webp":
            continue
        ws.append(list(row))

    for name in DEMOGRAPH_GALLERY:
        ws.append([
            "ПРОЕКТ — Демография ДВ",
            "ФОТОГАЛЕРЕЯ",
            "Фото с конференции",
            f"assets/photo/projects/demograph/{name}.webp",
        ])

    widths = [36, 22, 48, 52]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    wb.save(OUT)
    print(f"Saved {ws.max_row - 1} rows to {OUT}")


if __name__ == "__main__":
    main()
